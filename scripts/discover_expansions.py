"""
Discovery automatica degli expansion_id CardTrader per i fogli di MAIN.xlsx.

Per ogni foglio:
  - legge il "Card set name" in cella D3 (formato standard di MAIN.xlsx)
  - se non c'e', usa il nome del foglio normalizzato
  - matcha fuzzy contro la lista expansions di game_id=5 (Pokemon)
  - propone l'expansion_id con il punteggio di confidence

Output: expansions.json nella root del progetto (sovrascritto). I match
incerti hanno valore null e l'utente deve risolverli a mano.

USO:
  set CT_AUTH_TOKEN=eyJ...
  python scripts/discover_expansions.py [MAIN.xlsx]
"""

from __future__ import annotations
import argparse
import json
import sys
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from cardtrader_client import _get
from paths import EXPANSIONS_JSON, DEFAULT_WORKBOOK

# Fogli che NON sono espansioni: indici, note, raggruppamenti
EXCLUDED_SHEETS = {
    "NOTES_AND_TO_DO_LIST",
    "INDEX_TCG",
    "EX-SERIES",   # raggruppamento di tutte le EX
    "VS_ERA",      # raggruppamento delle VS
    "PROMO",       # raggruppamento di promo varie
}


def fetch_pokemon_expansions() -> list[dict]:
    """Scarica tutte le espansioni con game_id=5 (Pokemon)."""
    print("Carico espansioni Pokemon da CardTrader API...")
    data = _get("/expansions")
    pokemon = [e for e in data if e.get("game_id") == 5]
    print(f"  {len(pokemon)} espansioni Pokemon disponibili")
    return pokemon


def read_set_name(wb, sheet_name: str) -> str:
    """Legge il 'Card set name' da D3, o ripiega sul nome del foglio normalizzato."""
    s = wb[sheet_name]
    try:
        d3 = s["D3"].value
        if d3 and isinstance(d3, str) and d3.strip():
            return d3.strip()
    except Exception:
        pass
    return sheet_name.replace("_", " ").title()


def similarity(a: str, b: str) -> float:
    """Punteggio 0-100 di similarita' tra due stringhe (normalizzate)."""
    def norm(s):
        return "".join(c.lower() for c in s if c.isalnum())
    return SequenceMatcher(None, norm(a), norm(b)).ratio() * 100


def best_match(set_name: str, expansions: list[dict]) -> tuple[dict | None, float]:
    """Trova l'espansione con punteggio piu' alto. Ritorna (best, score)."""
    best = None
    best_score = 0.0
    for e in expansions:
        score = max(
            similarity(set_name, e.get("name", "")),
            similarity(set_name, e.get("code", "")),
        )
        if score > best_score:
            best_score = score
            best = e
    return best, best_score


def main():
    parser = argparse.ArgumentParser(description="Discovery automatica expansion_id")
    parser.add_argument("workbook", nargs="?", default=str(DEFAULT_WORKBOOK),
                        help=f"Percorso a MAIN.xlsx (default: {DEFAULT_WORKBOOK})")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).resolve()
    if not workbook_path.exists():
        print(f"ERRORE: {workbook_path} non trovato")
        sys.exit(1)

    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    pokemon_expansions = fetch_pokemon_expansions()

    result = {}
    uncertain = []
    not_found = []

    for sheet_name in wb.sheetnames:
        if sheet_name in EXCLUDED_SHEETS:
            print(f"[SKIP] {sheet_name}: foglio escluso")
            continue
        if wb[sheet_name].max_row < 6:
            print(f"[SKIP] {sheet_name}: foglio vuoto o senza dati")
            continue

        set_name = read_set_name(wb, sheet_name)
        match, score = best_match(set_name, pokemon_expansions)

        if score >= 80 and match:
            result[sheet_name] = match["id"]
            print(f"[OK]   {sheet_name:<35} '{set_name}' -> {match['name']} (id={match['id']}, score={score:.0f})")
        elif score >= 60 and match:
            result[sheet_name] = match["id"]
            uncertain.append((sheet_name, set_name, match, score))
            print(f"[?]    {sheet_name:<35} '{set_name}' -> {match['name']} (id={match['id']}, score={score:.0f})  ## VERIFICA")
        else:
            result[sheet_name] = None
            not_found.append((sheet_name, set_name, match, score))
            note = f"best={match['name']} ({score:.0f})" if match else "no match"
            print(f"[NO]   {sheet_name:<35} '{set_name}' -> {note}")

    EXPANSIONS_JSON.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"\n=== Riepilogo ===")
    print(f"  Match certi (>=80%): {sum(1 for v in result.values() if v) - len(uncertain)}")
    print(f"  Da verificare (60-80%): {len(uncertain)}")
    print(f"  Non trovati: {len(not_found)}")
    print(f"\nFile scritto: {EXPANSIONS_JSON}")

    if uncertain:
        print("\nDa verificare manualmente in expansions.json:")
        for s, n, m, sc in uncertain:
            print(f"  - {s}: '{n}' matcha '{m['name']}' (score={sc:.0f}, id={m['id']})")
    if not_found:
        print("\nNon trovati (impostati a null, vanno risolti a mano):")
        for s, n, m, sc in not_found:
            print(f"  - {s}: '{n}'")
        print("  Suggerimento: cerca a mano con:")
        print('    curl -s "https://api.cardtrader.com/api/v2/expansions" -H "Authorization: Bearer %CT_AUTH_TOKEN%" | jq \'.[] | select(.game_id==5) | select(.name | test("PAROLA_CHIAVE"; "i"))\'')


if __name__ == "__main__":
    main()
